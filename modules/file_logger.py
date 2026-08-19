"""
Maintains a running Excel log file for complaints (and leads), so the team
always has a single file with the full history — in addition to MongoDB.
Every time a new complaint/lead is saved, a row is appended to the relevant
file, and the updated file is emailed to the team as an attachment.
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
COMPLAINTS_LOG_PATH = os.path.join(DATA_DIR, "complaints_log.xlsx")
LEADS_LOG_PATH = os.path.join(DATA_DIR, "leads_log.xlsx")

COMPLAINT_HEADERS = [
    "Complaint ID", "Date/Time", "Mall", "Caller Name", "Phone",
    "Issue", "Priority", "Status",
]
LEAD_HEADERS = [
    "Lead ID", "Date/Time", "Mall", "Shop Number", "Full Name", "Phone",
    "Email", "Address", "Business Name", "Business Type", "Registered", "Status",
]


def _ensure_file(path, headers):
    """Creates the Excel file with a header row if it doesn't exist yet."""
    if not os.path.exists(path):
        os.makedirs(DATA_DIR, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = 22
        wb.save(path)


def log_complaint_to_file(complaint_data, complaint_id):
    """
    Appends one row to complaints_log.xlsx for this complaint.
    Returns the file path so it can be emailed as an attachment.
    """
    _ensure_file(COMPLAINTS_LOG_PATH, COMPLAINT_HEADERS)
    wb = load_workbook(COMPLAINTS_LOG_PATH)
    ws = wb["Log"]
    ws.append([
        str(complaint_id),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        complaint_data.get("mall_id"),
        complaint_data.get("full_name"),
        complaint_data.get("phone"),
        complaint_data.get("issue"),
        "Urgent" if complaint_data.get("is_urgent") else "Standard",
        "Open",
    ])
    wb.save(COMPLAINTS_LOG_PATH)
    return COMPLAINTS_LOG_PATH


def log_lead_to_file(lead_data, lead_id):
    """
    Appends one row to leads_log.xlsx for this lead.
    Returns the file path so it can be emailed as an attachment.
    """
    _ensure_file(LEADS_LOG_PATH, LEAD_HEADERS)
    wb = load_workbook(LEADS_LOG_PATH)
    ws = wb["Log"]
    ws.append([
        str(lead_id),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        lead_data.get("mall_id"),
        lead_data.get("shop_number"),
        lead_data.get("full_name"),
        lead_data.get("phone"),
        lead_data.get("email"),
        lead_data.get("address"),
        lead_data.get("business_name"),
        lead_data.get("business_type"),
        lead_data.get("is_registered"),
        "New",
    ])
    wb.save(LEADS_LOG_PATH)
    return LEADS_LOG_PATH
