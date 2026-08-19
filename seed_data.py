"""
One-time script to load A5 Mall's info and shop inventory into MongoDB,
reading directly from the demo Excel workbook (A5_Mall_Shop_Inventory_Demo.xlsx).

Usage:
    1. Place A5_Mall_Shop_Inventory_Demo.xlsx in the `data/` folder of this
       project (same structure as the demo file: "Mall Info" sheet +
       "Shop Inventory - A5 Mall" sheet).
    2. Run: python seed_data.py
    3. Re-run any time the Excel file is updated — it's safe to run multiple
       times (mall info is upserted, shop records are cleared and reloaded
       for this mall_id each run).

Requires: pip install openpyxl
"""
import os
from openpyxl import load_workbook
from modules.mongodb import seed_mall, seed_shops, shops_collection

MALL_ID = "A5"
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "A5_Mall_Shop_Inventory_Demo.xlsx")

# Maps the human-readable labels in the "Mall Info" sheet to the field names
# used elsewhere in the app (get_general_mall_info tool, etc.)
MALL_INFO_FIELD_MAP = {
    "Mall Name": "name",
    "City / Area": "city",
    "Full Address": "address",
    "Opening Hours (Sat–Thu)": "hours_sat_thu",
    "Opening Hours (Friday)": "hours_friday",
    "Parking": "parking",
    "Prayer Room": "prayer_room",
    "Leasing Office Phone": "leasing_phone",
    "Leasing Office Email": "leasing_email",
    "Duty Manager Phone (Escalations)": "duty_manager_phone",
    "Customer Service Number": "customer_service_phone",
    "Nearby Landmarks": "landmarks",
    "Mall Status": "status",
}


def read_mall_info(wb):
    """Reads the 'Mall Info' sheet (Field / Details two-column table) into a dict."""
    ws = wb["Mall Info"]
    mall_data = {"mall_id": MALL_ID}

    reading_rows = False
    for row in ws.iter_rows(values_only=True):
        field, value = row[0], row[1]
        if field == "Field" and value == "Details":
            reading_rows = True
            continue
        if not reading_rows or field is None:
            continue
        key = MALL_INFO_FIELD_MAP.get(field)
        if key:
            mall_data[key] = value

    return mall_data


def read_shop_inventory(wb):
    """Reads the 'Shop Inventory - A5 Mall' sheet into a list of shop dicts."""
    ws = wb["Shop Inventory - A5 Mall"]
    shops = []

    header_row_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Shop No.":
            header_row_found = True
            continue
        if not header_row_found or row[0] is None:
            continue

        shop_no, floor, size_sqft, category, status, tenant_type, rent_kwd, notes = row[:8]
        shops.append({
            "mall_id": MALL_ID,
            "shop_number": shop_no,
            "floor": floor,
            "size_sqft": size_sqft,
            "category": category,
            "status": status,           # "Available" or "Rented"
            "current_tenant_type": tenant_type if tenant_type != "-" else None,
            "rent_kwd": rent_kwd,
            "notes": notes,
        })

    return shops


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        print("Place A5_Mall_Shop_Inventory_Demo.xlsx inside a 'data/' folder next to seed_data.py.")
        return

    wb = load_workbook(EXCEL_PATH, data_only=True)

    # 1. Mall info
    mall_data = read_mall_info(wb)
    seed_mall(mall_data)
    print(f"Mall info seeded for {mall_data.get('name', MALL_ID)}.")

    # 2. Shop inventory — clear existing records for this mall first, so
    #    re-running the script doesn't create duplicates.
    shops_collection.delete_many({"mall_id": MALL_ID})
    shops = read_shop_inventory(wb)
    count = seed_shops(shops)
    print(f"{count} shops seeded for mall {MALL_ID}.")


if __name__ == "__main__":
    main()
